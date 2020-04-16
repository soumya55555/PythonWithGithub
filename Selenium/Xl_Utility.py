import openpyxl

def getRow_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.max_row

def getColumn_count(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column
def read_data(file,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=colnum).value
def write_data(file,sheetName,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=colnum).value=data
    workbook.save(file)