import openpyxl
# path="C:\Soumya\data.xlsx"
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active                       #workbook.get_sheet_by_name("Sheet1")
# rows=sheet.max_row
# colm=sheet.max_column
# print(rows)
# print(colm)
# for row in range(1,rows+1):
#     for col in range(1,colm+1):
#         print(sheet.cell(row= row,column= col).value , end="   ")
#
#     print( )


# writing in excel
path="C:\Soumya\data2.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value="welcome"

workbook.save(path)